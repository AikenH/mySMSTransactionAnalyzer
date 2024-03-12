import csv
import os
from utils.util import log_execution, setup_logger, is_number
from collections import defaultdict
import matplotlib.pyplot as plt
from datetime import datetime
import numpy as np
import seaborn as sns
import openpyxl

logger = setup_logger()
thin_border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(border_style='thin'),
                                             right=openpyxl.styles.borders.Side(border_style='thin'),
                                             top=openpyxl.styles.borders.Side(border_style='thin'),
                                             bottom=openpyxl.styles.borders.Side(border_style='thin'))
Font = openpyxl.styles.Font(u'微软雅黑', size=12, bold=False, italic=False, strike=False, color='000000')
ErrorFont = openpyxl.styles.Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color='FF0000')
RightFont = openpyxl.styles.Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color='008000')
Align = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)


@log_execution(verbose=False)
def generate_csv_files(transactions, output_dir):
    """
    Generates CSV files for transactions, grouped by account number.
    Each account's transactions are written to a separate CSV file.

    Args:
        transactions (list): The list of transaction details.p
        output_dir (str): The directory where the CSV files will be saved.
    """
    transactions_by_account = defaultdict(list)
    
    for transaction in transactions:
        account_number = transaction['account_number']
        transactions_by_account[account_number].append(transaction)

    for account_number, account_transactions in transactions_by_account.items():
        file_path = os.path.join(output_dir, f'账户{account_number}.xlsx')
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            headers = ['日期', '转出方', "接收方", "我方账号", "收入/支出", "金额", "余额", "银行名称", "局部预期余额计算", "差额(同前向)", "全局预期余额计算"]
            header_lens = [14, 36, 18, 12, 13, 12, 12, 14, 34, 14, 18]
            sheet.freeze_panes = 'A2'
            sheet.append(headers)

            for i, header in enumerate(headers, start=1):
                col_letter = openpyxl.utils.get_column_letter(i)
                cell = sheet.cell(row=1, column=i)
                cell.font = openpyxl.styles.Font(size=12, bold=True)
                cell.border = thin_border
                cell.alignment = Align
                
                cell.fill = openpyxl.styles.PatternFill(start_color='ffeb9c', end_color='ffeb9c', fill_type='solid')
                header_length = max(header_lens[i-1], 8)  # Minimum width of 10 characters
                sheet.column_dimensions[col_letter].width = header_length

            for idx, transaction in enumerate(account_transactions, start=2):
                row = list(transaction.values())
                row = [float(item) if is_number(item) else item for item in row]
                sheet.append(row)
                style_transaction_cell(sheet, idx, row)

            wb.save(file_path)

        except IOError as e:
            logger.error(f"IO error occurred while writing to {file_path}: {e}")
        except Exception as e:
            logger.error(f"An unexpected error occurred: {e}")


def style_transaction_cell(sheet, row, row_value):
    """
    Style a cell in the transaction sheet based on the transaction details.
    if the gap rol is not '', make font red when <0, green when >0
    if the note rol is not '', make font red.
    """
    map_idx_to_col = {'date': 0, 'object1': 1, 'object2': 2, 'account_number': 3, 'type': 4,
                      'amount': 5, 'balance': 6, 'bank_name': 7, 'note': 8, 'gap': 9, 'running_balance': 10}
    cell = sheet.cell(row=row, column=1)

    for col_idx, col_value in enumerate(row_value, start=0):
        cell = sheet.cell(row=row, column=col_idx+1)
        cell.border = thin_border
        cell.alignment = Align
        if col_idx == map_idx_to_col['gap']:
            if col_value == '':
                continue
            elif col_value < 0:
                # cell.fill = openpyxl.styles.PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')
                cell.font = ErrorFont

            else:
                # cell.fill = openpyxl.styles.PatternFill(start_color='ccffcc', end_color='ccffcc', fill_type='solid')
                cell.font = RightFont
        elif col_idx == map_idx_to_col['note']:
            if col_value == '' or '没有余额' in col_value:
                continue
            else:
                # cell.fill = openpyxl.styles.PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')
                cell.font = ErrorFont
        else:
            continue

    return


@log_execution(verbose=False)
def calculate_monthly_totals(transactions):
    """Calculate monthly totals for income and outcome, grouped by account number."""
    monthly_totals = defaultdict(lambda: defaultdict(lambda: {'income': 0, 'outcome': 0}))

    for transaction in transactions:
        date = datetime.strptime(transaction['date'], '%Y-%m-%d')
        amount = float(transaction['amount'])  # Convert to float or int
        account_number = transaction['account_number']

        if amount > 0:
            monthly_totals[account_number][date.strftime('%Y-%m')]['income'] += amount
        else:
            monthly_totals[account_number][date.strftime('%Y-%m')]['outcome'] += abs(amount)

    return monthly_totals


# Setting Seaborn style
sns.set(style="whitegrid")


@log_execution(verbose=False)
def plot_monthly_totals(monthly_totals, output_dir):
    """Plot the monthly totals as side-by-side bar charts with improved styling."""
    for account, months in monthly_totals.items():
        fig, ax = plt.subplots()
        months_sorted = sorted(months.keys())
        incomes = [months[month]['income'] for month in months_sorted]
        outcomes = [months[month]['outcome'] for month in months_sorted]

        bar_width = 0.35
        index = np.arange(len(months_sorted))

        bar1 = ax.bar(index, incomes, bar_width, label='Income', color='skyblue')
        bar2 = ax.bar(index + bar_width, outcomes, bar_width, label='Outcome', color='salmon')

        ax.set_xlabel('Month', fontsize=10, fontweight='bold')
        ax.set_ylabel('Amount', fontsize=12, fontweight='bold')
        ax.set_title(f'Monthly Income and Outcome Comparison for Account {account}', fontsize=14, fontweight='bold')

        month_labels = [month[:4] + '-' + month[5:] if i != 0 else month for i, month in enumerate(months_sorted)]
        ax.set_xticks(index + bar_width / 2)
        ax.set_xticklabels(month_labels, rotation=0, fontsize=10, ha='center')

        # Add padding to X- and Y-axes
        ax.set_ylim(0, max(incomes + outcomes) * 1.1)  # 10% padding above the max value
        ax.margins(x=0.1)  # 5% padding for the x-axis

        ax.legend()

        # Adding grid lines
        ax.yaxis.grid(True)

        def autolabel(bars):
            """Attach a text label above each bar in *bars*, displaying its height."""
            for bar in bars:
                height = bar.get_height()
                ax.annotate('{}'.format(round(height, 2)),
                            xy=(bar.get_x() + bar.get_width() / 2, height),
                            xytext=(0, 3),  # 3 points vertical offset
                            textcoords="offset points",
                            ha='center', va='bottom', fontsize=9)

        autolabel(bar1)
        autolabel(bar2)

        plt.tight_layout()
        # plt.show()

        # save the plot
        file_name = f'{account}.png'
        file_path = os.path.join(output_dir, file_name)
        plt.savefig(file_path, dpi=300)
        plt.close()
